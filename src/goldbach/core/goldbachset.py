import os
from collections import OrderedDict as odict
from openpyxl import Workbook
from goldbach.core.rationalset import RationalSet

basedir = r'C:\Users\enrique\Google Drive\Enrique\Goldbach\hojas'


class Goldbach(object):
	"""Holder for rationals goldbach paris for an even number

	Args:
		number (int): even number to check goldbach pairs
	"""
	def __init__(self, number):
		self.number = number
		self.pairs = []

		for a in range(1, self.number / 2 + 1, 2):
			b = self.number - a
			add = a + b
			prod = a * b
			result = prod - add + 1
			pair = {
				'a': a,
				'b': b,
				'a+b': add,
				'a*b': prod,
				'result': result,
				'periods': odict()
			}
			for base in range(2, 13):
				ra = RationalSet(a, base)
				rb = RationalSet(b, base)
				rprod = RationalSet(prod, base)
				rprodperiod = rprod.getPeriod()
				div = result / float(rprodperiod) if rprodperiod else 0
				period = {
					'a': ra.getPeriod(),
					'b': rb.getPeriod(),
					'a*b': rprodperiod,
					'div': div
				}
				pair['periods'][base] = period
			self.pairs.append(pair)

	def output(self, ws, row):
		"""Output goldbach pairs to an excel sheet

		Args:
			name (str): file name
		"""
		for pair in self.pairs:
			ws.cell(row=row, column=1, value=pair['a'])
			ws.cell(row=row, column=2, value=pair['b'])
			ws.cell(row=row, column=3, value=pair['a+b'])
			ws.cell(row=row, column=4, value=pair['a*b'])
			ws.cell(row=row, column=5, value=pair['result'])
			column = 6
			for period in pair['periods'].values():
				ws.cell(row=row, column=column, value=period['a'] or '')
				ws.cell(row=row, column=column + 1, value=period['b'] or '')
				ws.cell(row=row, column=column + 2, value=period['a*b'] or '')
				ws.cell(row=row, column=column + 3, value=period['div'] or '')
				column += 4
			row += 1
		return row


def outNumbers(name):
	wb = Workbook()
	ws = wb.active

	row = 1
	for p in range(2, 13):
		column = 6 + 4 * (p - 1)
		ws.cell(row=row, column=column, value='base {0}'.format(p))

	row = 2
	ws.cell(row=row, column=1, value='a')
	ws.cell(row=row, column=2, value='b')
	ws.cell(row=row, column=3, value='a+b')
	ws.cell(row=row, column=4, value='a*b')
	ws.cell(row=row, column=5, value='result')
	for base in range(2, 13):
		column = 6 + 4 * (base - 2)
		ws.cell(row=row, column=column, value='a')
		ws.cell(row=row, column=column + 1, value='b')
		ws.cell(row=row, column=column + 2, value='a*b')
		ws.cell(row=row, column=column + 3, value='div')

	row = 3
	for number in range(4, 24, 2):
		goldbach = Goldbach(number)
		row = goldbach.output(ws, row) + 1

	wb.save(os.path.join(basedir, name))


if __name__ == '__main__':
	outNumbers('goldbach_4_24.xlsx')
